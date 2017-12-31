import sys
import io
import os
import csv
import json
from .models import *

try:
    from openpyxl import *

    def csvtoxlsx(csv_name, xlsx_name, floats=None, csv_dialect='excel-tab', quoting=csv.QUOTE_ALL):
        """
        A function to convert a CSV file to XLSX so it can be used by openpyxl.
        csvname = path of the csv you want to convert (include .csv)
        xlsx_name = path of the resulting xlsx file (include .xlsx)
        floats = A list of column indexes in which numbers appear
        """

        f = open(csv_name, 'rt')
        reader = csv.reader(f, dialect=csv_dialect, quoting=csv.QUOTE_ALL)
        wb = Workbook()
        dest_filename = xlsx_name
        ws = wb.worksheets[0]
        ws.title = os.path.basename(os.path.splitext(xlsx_name)[
                                    0])  # chop off extension
        if floats is None:
            floats = []
        for row_index, row in enumerate(reader):
            for column_index, cell in enumerate(row):
                if column_index in floats:
                    s = cell
                    # Handles heading row or non floats
                    try:
                        s = float(s)
                        ws.cell(column=column_index + 1,
                                row=row_index + 1).value = s

                    except ValueError:
                        ws.cell(column=column_index + 1,
                                row=row_index + 1).value = s

                elif column_index not in floats:
                    # Handles openpyxl 'illegal chars'
                    try:
                        ws.cell(column=column_index + 1,
                                row=row_index + 1).value = cell

                    except:
                        ws.cell(column=column_index + 1,
                                row=row_index + 1).value = 'illegal char'

        wb.save(filename=dest_filename)
except ImportError:
    pass


class RegistrationData:

    def load_from_db(self):
        pass

    def dump_data(self):
        raw_data = [{'serial_no': i, 'entry': x.__dict__}
                    for i, x in enumerate(self.load_from_db())]
        data = {}
        data['total_entries'] = len(raw_data)
        data['entries'] = raw_data
        return data

    def prettify(self, item):
        return str(item).replace('_', ' ').title()

    def dump_to_csv(self):
        data = self.dump_data()
        output = io.StringIO()
        csv_writer = csv.writer(
            output, dialect='excel-tab', quoting=csv.QUOTE_ALL)
        csv_writer.writerow(["Total Entries", data['total_entries']])
        for entry in data['entries']:
            csv_writer.writerow(["Serial No.", entry['serial_no']])
            csv_writer.writerow(["Entry"])
            entry_data = entry['entry']

            entry_without_sub_entries = {
                self.prettify(key): value
                for key, value in entry_data.items()
                if not isinstance(value, list)}
            csv_writer.writerow(
                list(entry_without_sub_entries.keys()))
            csv_writer.writerow(
                list(entry_without_sub_entries.values()))

            entry_with_only_sub_entries = {
                self.prettify(key): value
                for key, value in entry_data.items()
                if isinstance(value, list)}
            for entry_type, entry in entry_with_only_sub_entries.items():
                csv_writer.writerow([entry_type])
                for sub_entry in entry:
                    csv_writer.writerow((self.prettify(key)
                                         for key in sub_entry.keys()))
                    csv_writer.writerow(sub_entry.values())
        final_output = output.getvalue()
        output.close()
        return final_output

    def dump_to_json(self):
        return json.dumps(self.dump_data(), indent=4)

    def save_all_data_to_file(self, path, json=False):
        with open(path, 'w') as output:
            self.save_all_data(json=json, file=output)

    def save_all_data(self, json=False, file=sys.stdout):
        print(self.dump_to_json() if json else self.dump_to_csv(), file=file)


class ProsceniumTheatreRegistrationData(RegistrationData):

    def load_from_db(self):
        for registration in ProsceniumTheatreRegistration.objects.all():
            data = ProsceniumTheatreRegistrationData()
            data.time = str(registration.time)
            data.id = registration.id
            data.referral_code = registration.referral_code
            data.institution = registration.institution
            data.language = registration.language
            data.contact1 = registration.contact1
            data.contact2 = registration.contact2
            data.email = registration.email
            data.prelims_video = str(registration.prelims_video)
            data.prelims_script = str(registration.prelims_script)
            data.performers = []
            data.accompanists = []
            for participant in registration.prosceniumtheatreparticipant_set.all():
                participant_data = {}
                participant_data['name'] = participant.name
                participant_data['age'] = participant.age
                participant_data['photo'] = str(participant.photo)
                if participant.role == ProsceniumParticipant.PERFORMER:
                    data.performers.append(participant_data)
                elif participant.role == ProsceniumParticipant.ACCOMPANIST:
                    data.accompanists.append(participant_data)
            data.total_performers = len(data.performers)
            data.total_accompanists = len(data.accompanists)
            yield data


class ProsceniumStreetPlayRegistrationData(RegistrationData):

    def load_from_db(self):
        for registration in ProsceniumStreetPlayRegistration.objects.all():
            data = ProsceniumStreetPlayRegistrationData()
            data.time = str(registration.time)
            data.id = registration.id
            data.referral_code = registration.referral_code
            data.institution = registration.institution
            data.language = registration.language
            data.contact1 = registration.contact1
            data.contact2 = registration.contact2
            data.email = registration.email
            data.performers = []
            data.accompanists = []
            for participant in registration.prosceniumstreetplayparticipant_set.all():
                participant_data = {}
                participant_data['name'] = participant.name
                participant_data['age'] = participant.age
                participant_data['photo'] = str(participant.photo)
                if participant.role == ProsceniumParticipant.PERFORMER:
                    data.performers.append(participant_data)
                elif participant.role == ProsceniumParticipant.ACCOMPANIST:
                    data.accompanists.append(participant_data)
            data.total_performers = len(data.performers)
            data.total_accompanists = len(data.accompanists)
            yield data


class BoBRegistrationData(RegistrationData):

    def load_from_db(self):
        for registration in BoBRegistration.objects.all():
            data = BoBRegistrationData()
            data.time = str(registration.time)
            data.id = registration.id
            data.referral_code = registration.referral_code
            data.band_name = registration.band_name
            data.city = registration.city
            data.genre = registration.genre
            data.facebook = registration.facebook_link
            data.email = registration.email
            data.prelims_venue = registration.prelims_venue
            data.audio_sample_file = str(registration.audio_sample_file)
            data.audio_sample_link = str(registration.audio_sample_link)
            data.participants = []
            for participant in registration.bobparticipant_set.all():
                participant_data = {}
                participant_data['name'] = participant.name
                participant_data['contact'] = participant.contact
                participant_data['instrument'] = participant.instrument
                data.participants.append(participant_data)
            data.participant_count = len(data.participants)
            yield data


class LasyaRegistrationData(RegistrationData):

    def load_from_db(self):
        for registration in LasyaRegistration.objects.all():
            data = LasyaRegistrationData()
            data.time = str(registration.time)
            data.id = registration.id
            data.referral_code = registration.referral_code
            data.name = registration.name
            data.institution = registration.institution
            data.contact = registration.contact
            data.email = registration.email
            data.prelims_video = str(registration.prelims_video)
            data.participants = []
            for participant in registration.lasyaparticipant_set.all():
                participant_data = {}
                participant_data['name'] = participant.name
                participant_data['contact'] = participant.contact
                participant_data['email'] = participant.email
                data.participants.append(participant_data)
            data.participant_count = len(data.participants)
            yield data


class SInECRegistrationData(RegistrationData):

    def load_from_db(self):
        for registration in SInECRegistration.objects.all():
            data = SInECRegistrationData()
            data.time = str(registration.time)
            data.id = registration.id
            data.referral_code = registration.referral_code
            data.team_name = registration.team_name
            data.project_name = registration.project_name
            data.project_field = registration.project_field
            data.project_abstract = registration.project_abstract
            data.project_patented = registration.project_patented
            data.registered_company = registration.registered_company
            data.privacy_preference = registration.privacy_preference
            data.address = registration.address
            data.email = registration.email
            data.contact = registration.contact
            data.project_file = str(registration.project_file)
            data.project_video = str(registration.project_video)
            data.project_video_link = str(registration.project_video_link)
            data.participants = []
            for participant in registration.sinecparticipant_set.all():
                participant_data = {}
                participant_data['name'] = participant.name
                participant_data['city'] = participant.city
                participant_data['institution'] = participant.institution
                participant_data['student_type'] = participant.student_type
                data.participants.append(participant_data)
            data.participant_count = len(data.participants)
            yield data


class OpenMicRegistrationData(RegistrationData):

    def load_from_db(self):
        for registration in OpenMicRegistration.objects.all():
            data = OpenMicRegistrationData()
            data.time = str(registration.time)
            data.id = registration.id
            data.referral_code = registration.referral_code
            data.name = registration.name
            data.email = registration.email
            data.event = registration.event
            data.expected_performance_duration_mins = registration.expected_performance_duration_mins
            data.instrument_requirement = registration.instrument_requirement
            data.reason_for_gt_3_members = registration.reason_for_gt_3_members
            data.participants = []
            for participant in registration.openmicparticipant_set.all():
                participant_data = {}
                participant_data['name'] = participant.name
                data.participants.append(participant_data)
        data.participant_count = len(data.participants)
        yield data


class DecoherenceRegistrationData(RegistrationData):

    def load_from_db(self):
        for registration in DecoherenceRegistration.objects.all():
            data = DecoherenceRegistrationData()
            data.time = str(registration.time)
            data.id = registration.id
            data.referral_code = registration.referral_code
            data.team_name = registration.team_name
            data.participants = []
            for participant in registration.decoherenceparticipant_set.all():
                participant_data = {}
                participant_data['name'] = participant.name
                participant_data['institution'] = participant.institution
                participant_data['contact'] = participant.contact
                participant_data['email'] = participant.email
                data.participants.append(participant_data)
        data.participant_count = len(data.participants)
        yield data

class HackathonRegistrationData(RegistrationData):

    def load_from_db(self):
        for registration in HackathonRegistration.objects.all():
            data = HackathonRegistrationData()
            data.time = str(registration.time)
            data.id = registration.id
            data.referral_code = registration.referral_code
            data.team_name = registration.team_name
            data.contact = registration.contact
            data.email = registration.email
            data.abstract = registration.abstract
            data.participants = []
            for participant in registration.Hackathonparticipant_set.all():
                participant_data = {}
                participant_data['name'] = participant.name
                data.participants.append(participant_data)
        data.participant_count = len(data.participants)
        yield data
